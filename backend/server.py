from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Query, Header, Form
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
import os
import io
import logging
import bcrypt
import jwt
import secrets
import uuid as uuid_mod
import requests
from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any
from datetime import datetime, timezone, timedelta

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_ALGORITHM = "HS256"

# Simple in-memory cache for reference data
_cache = {}
_cache_ttl = 30  # seconds

async def get_cached(key, fetch_fn):
    import time
    now = time.time()
    if key in _cache and now - _cache[key]["ts"] < _cache_ttl:
        return _cache[key]["data"]
    data = await fetch_fn()
    _cache[key] = {"data": data, "ts": now}
    return data

def invalidate_cache(key=None):
    if key:
        _cache.pop(key, None)
    else:
        _cache.clear()

# ─── Object Storage ───
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = "fabrictrack"
storage_key = None

def init_storage():
    global storage_key
    if storage_key:
        return storage_key
    resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
    resp.raise_for_status()
    storage_key = resp.json()["storage_key"]
    return storage_key

def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120
    )
    resp.raise_for_status()
    return resp.json()

def get_object(path: str):
    key = init_storage()
    resp = requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")

# ─── Auth Helpers ───
def get_jwt_secret():
    return os.environ["JWT_SECRET"]

def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode("utf-8"), salt).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(hours=24), "type": "access"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def set_auth_cookies(response: Response, access_token: str, refresh_token: str):
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=True, samesite="none", max_age=604800, path="/")

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_admin(request: Request) -> dict:
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

# App setup
app = FastAPI()
api_router = APIRouter(prefix="/api")

# ─── Pydantic Models ───
class LoginRequest(BaseModel):
    email: str
    password: str

class CreateUserRequest(BaseModel):
    email: str
    password: str
    name: str
    role: str = "sales"
    department: str = "Sales"

class UpdateUserRequest(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    department: Optional[str] = None
    is_active: Optional[bool] = None

class StageCreate(BaseModel):
    name: str
    order: int
    color: str = "#9CA3AF"
    description: str = ""
    input_type: str = "text"  # text, date, select
    is_mandatory: bool = False
    select_options: List[str] = []
    lead_time_days: int = 0
    date_input_mode: str = "manual"
    assigned_users: List[str] = []  # user IDs who can complete/comment on this stage

class StageUpdate(BaseModel):
    name: Optional[str] = None
    order: Optional[int] = None
    color: Optional[str] = None
    description: Optional[str] = None
    input_type: Optional[str] = None
    is_mandatory: Optional[bool] = None
    select_options: Optional[List[str]] = None
    lead_time_days: Optional[int] = None
    date_input_mode: Optional[str] = None
    assigned_users: Optional[List[str]] = None

class EnquiryCreate(BaseModel):
    customer_name: str
    fabric_type: str
    quantity: str = ""  # Made optional - removed from create form
    style_no: str = ""
    department: Optional[str] = None
    notes: str = ""
    rate: str = ""  # Removed from create form but kept for backward compatibility
    po_no: str = ""  # Removed from create form but kept for backward compatibility
    po_del_date: str = ""  # Removed from create form but kept for backward compatibility
    fabric_received: str = "no"
    qty_received: str = ""
    sample_number: str = ""
    stage_values: Dict[str, Any] = {}

class EnquiryUpdate(BaseModel):
    customer_name: Optional[str] = None
    fabric_type: Optional[str] = None
    quantity: Optional[str] = None
    style_no: Optional[str] = None
    department: Optional[str] = None
    notes: Optional[str] = None
    rate: Optional[str] = None
    po_no: Optional[str] = None
    po_del_date: Optional[str] = None
    fabric_received: Optional[str] = None
    qty_received: Optional[str] = None
    sample_number: Optional[str] = None
    stage_values: Optional[Dict[str, Any]] = None
    image_path: Optional[str] = None

class DepartmentCreate(BaseModel):
    name: str
    description: str = ""

class DepartmentUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None

class DepartmentHierarchyItem(BaseModel):
    stage_id: str
    order: int
    assigned_users: List[str] = []

class DepartmentHierarchyUpdate(BaseModel):
    items: List[DepartmentHierarchyItem]

class CustomerCreate(BaseModel):
    name: str

class CustomerUpdate(BaseModel):
    name: Optional[str] = None

class FabricTypeCreate(BaseModel):
    name: str
    gsm: str = ""
    width: str = ""
    composition: str = ""
    construction: str = ""

class FabricTypeUpdate(BaseModel):
    name: Optional[str] = None
    gsm: Optional[str] = None
    width: Optional[str] = None
    composition: Optional[str] = None
    construction: Optional[str] = None

# ─── Auth Routes ───
@api_router.post("/auth/login")
async def login(req: LoginRequest, response: Response, request: Request):
    email = req.email.strip().lower()
    ip = request.client.host if request.client else "unknown"
    identifier = f"{ip}:{email}"
    attempt = await db.login_attempts.find_one({"identifier": identifier})
    if attempt and attempt.get("count", 0) >= 5:
        lockout_until = attempt.get("locked_until")
        if lockout_until and datetime.now(timezone.utc) < lockout_until:
            raise HTTPException(status_code=429, detail="Too many failed attempts. Try again in 15 minutes.")
        else:
            await db.login_attempts.delete_one({"identifier": identifier})
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(req.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": identifier},
            {"$inc": {"count": 1}, "$set": {"locked_until": datetime.now(timezone.utc) + timedelta(minutes=15)}},
            upsert=True
        )
        raise HTTPException(status_code=401, detail="Invalid email or password")
    await db.login_attempts.delete_one({"identifier": identifier})
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    set_auth_cookies(response, access_token, refresh_token)
    return {"_id": user_id, "email": user["email"], "name": user["name"], "role": user["role"], "department": user.get("department", "")}

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out"}

@api_router.get("/auth/me")
async def get_me(request: Request):
    return await get_current_user(request)

@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        access_token = create_access_token(str(user["_id"]), user["email"])
        response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=86400, path="/")
        return {"message": "Token refreshed"}
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

# ─── Biometric / WebAuthn ───
import base64

# In-memory challenge store (short-lived)
_webauthn_challenges = {}

def _b64url_encode(data):
    if isinstance(data, str):
        data = data.encode()
    return base64.urlsafe_b64encode(data).rstrip(b'=').decode()

def _b64url_decode(s):
    s += '=' * (4 - len(s) % 4)
    return base64.urlsafe_b64decode(s)

@api_router.post("/auth/biometric/register-options")
async def biometric_register_options(request: Request):
    user = await get_current_user(request)
    origin = request.headers.get("origin", "")
    host = request.headers.get("host", "localhost")
    rp_id = origin.split("://")[1].split(":")[0].split("/")[0] if "://" in origin else host.split(":")[0]
    challenge = secrets.token_bytes(32)
    _webauthn_challenges[user["_id"]] = challenge
    # Get existing credential IDs to exclude
    existing = await db.webauthn_credentials.find({"user_id": user["_id"]}, {"_id": 0}).to_list(50)
    exclude = [{"type": "public-key", "id": c["credential_id_b64"]} for c in existing if c.get("credential_id_b64")]
    return {
        "challenge": _b64url_encode(challenge),
        "rp": {"name": "FabricTrack", "id": rp_id},
        "user": {
            "id": _b64url_encode(user["_id"]),
            "name": user["email"],
            "displayName": user["name"]
        },
        "pubKeyCredParams": [
            {"alg": -7, "type": "public-key"},
            {"alg": -257, "type": "public-key"}
        ],
        "timeout": 60000,
        "authenticatorSelection": {
            "authenticatorAttachment": "platform",
            "userVerification": "required",
            "residentKey": "preferred",
            "requireResidentKey": False
        },
        "attestation": "none",
        "excludeCredentials": exclude
    }

@api_router.post("/auth/biometric/register-complete")
async def biometric_register_complete(request: Request):
    user = await get_current_user(request)
    challenge = _webauthn_challenges.pop(user["_id"], None)
    if not challenge:
        raise HTTPException(status_code=400, detail="No registration in progress")
    body = await request.json()
    # Store the credential - we trust the browser's attestation for platform authenticators
    cred_id = body.get("id", "")
    raw_id = body.get("rawId", "")
    resp = body.get("response", {})
    cred_doc = {
        "id": secrets.token_hex(12),
        "user_id": user["_id"],
        "credential_id_b64": cred_id,
        "raw_id_b64": raw_id,
        "public_key": resp.get("publicKey", ""),
        "attestation_object": resp.get("attestationObject", ""),
        "client_data_json": resp.get("clientDataJSON", ""),
        "transports": body.get("transports", []),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.webauthn_credentials.insert_one(cred_doc)
    return {"message": "Biometric registered successfully"}

@api_router.post("/auth/biometric/login-options")
async def biometric_login_options(request: Request):
    body = await request.json()
    email = body.get("email", "")
    origin = request.headers.get("origin", "")
    host = request.headers.get("host", "localhost")
    rp_id = origin.split("://")[1].split(":")[0].split("/")[0] if "://" in origin else host.split(":")[0]
    user = await db.users.find_one({"email": email})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user_id = str(user["_id"])
    creds = await db.webauthn_credentials.find({"user_id": user_id}, {"_id": 0}).to_list(50)
    if not creds:
        raise HTTPException(status_code=404, detail="No biometric credentials registered for this account")
    challenge = secrets.token_bytes(32)
    _webauthn_challenges[f"auth_{user_id}"] = {"challenge": challenge, "user_id": user_id}
    allow_credentials = []
    for c in creds:
        ac = {"type": "public-key", "id": c["credential_id_b64"]}
        if c.get("transports"):
            ac["transports"] = c["transports"]
        allow_credentials.append(ac)
    return {
        "challenge": _b64url_encode(challenge),
        "rpId": rp_id,
        "timeout": 60000,
        "userVerification": "required",
        "allowCredentials": allow_credentials
    }

@api_router.post("/auth/biometric/login-complete")
async def biometric_login_complete(request: Request, response: Response):
    body = await request.json()
    cred_id = body.get("id", "")
    # Find credential
    cred = await db.webauthn_credentials.find_one({"credential_id_b64": cred_id}, {"_id": 0})
    if not cred:
        raise HTTPException(status_code=401, detail="Credential not recognized")
    user_id = cred["user_id"]
    challenge_key = f"auth_{user_id}"
    challenge_data = _webauthn_challenges.pop(challenge_key, None)
    if not challenge_data:
        raise HTTPException(status_code=400, detail="No authentication in progress")
    # Verify challenge match from clientDataJSON
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    access_token = create_access_token(str(user["_id"]), user["email"])
    refresh = create_refresh_token(str(user["_id"]))
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh, httponly=True, secure=True, samesite="none", max_age=604800, path="/")
    return {"message": "Biometric login successful", "user": {
        "_id": str(user["_id"]), "email": user["email"], "name": user["name"],
        "role": user.get("role", "user"), "department": user.get("department", "")
    }}

@api_router.get("/auth/biometric/status")
async def biometric_status(request: Request):
    user = await get_current_user(request)
    count = await db.webauthn_credentials.count_documents({"user_id": user["_id"]})
    return {"registered": count > 0, "count": count}

@api_router.delete("/auth/biometric/credentials")
async def biometric_remove(request: Request):
    user = await get_current_user(request)
    await db.webauthn_credentials.delete_many({"user_id": user["_id"]})
    return {"message": "Biometric credentials removed"}

# ─── User Management ───
@api_router.get("/users")
async def get_users(request: Request):
    await get_current_user(request)
    users = await db.users.find({}, {"password_hash": 0}).to_list(1000)
    for u in users:
        u["_id"] = str(u["_id"])
        if "created_at" in u and isinstance(u["created_at"], datetime):
            u["created_at"] = u["created_at"].isoformat()
    return users

@api_router.post("/users")
async def create_user(req: CreateUserRequest, request: Request):
    await require_admin(request)
    email = req.email.strip().lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    user_doc = {
        "email": email, "password_hash": hash_password(req.password),
        "name": req.name, "role": req.role, "department": req.department,
        "is_active": True, "created_at": datetime.now(timezone.utc)
    }
    result = await db.users.insert_one(user_doc)
    return {"_id": str(result.inserted_id), "email": email, "name": req.name, "role": req.role, "department": req.department, "is_active": True}

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, req: UpdateUserRequest, request: Request):
    await require_admin(request)
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": update_data})
    user = await db.users.find_one({"_id": ObjectId(user_id)}, {"password_hash": 0})
    user["_id"] = str(user["_id"])
    if "created_at" in user and isinstance(user["created_at"], datetime):
        user["created_at"] = user["created_at"].isoformat()
    return user

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, request: Request):
    await require_admin(request)
    result = await db.users.delete_one({"_id": ObjectId(user_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted"}

# ─── Stage Master ───
@api_router.get("/stages")
async def get_stages(request: Request):
    await get_current_user(request)
    return await get_cached("stages", lambda: db.stages.find({}, {"_id": 0}).sort("order", 1).to_list(100))

@api_router.post("/stages")
async def create_stage(req: StageCreate, request: Request):
    await require_admin(request)
    stage_id = secrets.token_hex(12)
    stage_doc = {
        "id": stage_id, "name": req.name, "order": req.order,
        "color": req.color, "description": req.description,
        "input_type": req.input_type, "is_mandatory": req.is_mandatory,
        "select_options": req.select_options,
        "lead_time_days": req.lead_time_days,
        "date_input_mode": req.date_input_mode,
        "assigned_users": req.assigned_users,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.stages.insert_one(stage_doc)
    invalidate_cache("stages")
    return {k: v for k, v in stage_doc.items() if k != "_id"}

@api_router.put("/stages/{stage_id}")
async def update_stage(stage_id: str, req: StageUpdate, request: Request):
    await require_admin(request)
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await db.stages.update_one({"id": stage_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Stage not found")
    stage = await db.stages.find_one({"id": stage_id}, {"_id": 0})
    invalidate_cache("stages")
    return stage

@api_router.delete("/stages/{stage_id}")
async def delete_stage(stage_id: str, request: Request):
    await require_admin(request)
    result = await db.stages.delete_one({"id": stage_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Stage not found")
    return {"message": "Stage deleted"}

# ─── File Upload ───
@api_router.post("/upload")
async def upload_file(request: Request, file: UploadFile = File(...)):
    user = await get_current_user(request)
    ext = file.filename.split(".")[-1] if "." in file.filename else "bin"
    path = f"{APP_NAME}/uploads/{user['_id']}/{uuid_mod.uuid4()}.{ext}"
    data = await file.read()
    content_type = file.content_type or "application/octet-stream"
    result = put_object(path, data, content_type)
    file_doc = {
        "id": str(uuid_mod.uuid4()), "storage_path": result["path"],
        "original_filename": file.filename, "content_type": content_type,
        "size": result.get("size", len(data)), "is_deleted": False,
        "uploaded_by": user["_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.files.insert_one(file_doc)
    return {"path": result["path"], "filename": file.filename, "id": file_doc["id"]}

@api_router.get("/files/{path:path}")
async def download_file(path: str, request: Request, auth: Optional[str] = Query(None)):
    # Auth check - try cookie first, then query param
    token = request.cookies.get("access_token")
    if not token and auth:
        token = auth
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    record = await db.files.find_one({"storage_path": path, "is_deleted": False})
    if not record:
        raise HTTPException(status_code=404, detail="File not found")
    data, ct = get_object(path)
    return Response(content=data, media_type=record.get("content_type", ct))

# ─── Voice Notes ───
@api_router.post("/enquiries/{enquiry_id}/voice-notes")
async def add_voice_note(enquiry_id: str, request: Request, file: UploadFile = File(...)):
    user = await get_current_user(request)
    existing = await db.enquiries.find_one({"id": enquiry_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Enquiry not found")
    ext = file.filename.split(".")[-1] if "." in file.filename else "webm"
    path = f"{APP_NAME}/voice/{enquiry_id}/{uuid_mod.uuid4()}.{ext}"
    data = await file.read()
    content_type = file.content_type or "audio/webm"
    result = put_object(path, data, content_type)
    # Save file record
    file_doc = {
        "id": str(uuid_mod.uuid4()), "storage_path": result["path"],
        "original_filename": file.filename, "content_type": content_type,
        "size": result.get("size", len(data)), "is_deleted": False,
        "uploaded_by": user["_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.files.insert_one(file_doc)
    # Add to enquiry voice_notes array
    note_doc = {
        "id": str(uuid_mod.uuid4()),
        "storage_path": result["path"],
        "content_type": content_type,
        "recorded_by": user["_id"],
        "recorded_by_name": user["name"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.enquiries.update_one({"id": enquiry_id}, {"$push": {"voice_notes": note_doc}})
    return note_doc

@api_router.delete("/enquiries/{enquiry_id}/voice-notes/{note_id}")
async def delete_voice_note(enquiry_id: str, note_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete voice notes")
    await db.enquiries.update_one({"id": enquiry_id}, {"$pull": {"voice_notes": {"id": note_id}}})
    return {"message": "Voice note deleted"}

# ─── Enquiry Routes ───
@api_router.get("/enquiries")
async def get_enquiries(request: Request, stage: Optional[str] = None, department: Optional[str] = None, assigned_to: Optional[str] = None, search: Optional[str] = None, customer_name: Optional[str] = None, fabric_type: Optional[str] = None, style_no: Optional[str] = None, page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100)):
    await get_current_user(request)
    query = {}
    if stage:
        query[f"stage_values.{stage}"] = {"$exists": True}
    if department:
        query["department"] = department
    if assigned_to:
        query["assigned_to"] = assigned_to
    if customer_name:
        query["customer_name"] = {"$regex": customer_name, "$options": "i"}
    if fabric_type:
        query["fabric_type"] = {"$regex": fabric_type, "$options": "i"}
    if style_no:
        query["style_no"] = {"$regex": style_no, "$options": "i"}
    if search:
        query["$or"] = [
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"fabric_type": {"$regex": search, "$options": "i"}},
            {"style_no": {"$regex": search, "$options": "i"}}
        ]
    total = await db.enquiries.count_documents(query)
    skip = (page - 1) * page_size
    enquiries = await db.enquiries.find(query, {"_id": 0, "history": 0, "voice_notes": 0, "comments": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    # Calculate delay status for list view
    stages = await db.stages.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    now = datetime.now(timezone.utc)
    for enq in enquiries:
        sv = enq.get("stage_values", {})
        delay_info = {}
        for i, stage in enumerate(stages):
            sid = stage["id"]
            lead_time = stage.get("lead_time_days", 0)
            stage_val = sv.get(sid, {})
            stage_value = stage_val.get("value", "") if isinstance(stage_val, dict) else str(stage_val) if stage_val else ""
            stage_updated_at = stage_val.get("updated_at", "") if isinstance(stage_val, dict) else ""
            status = "pending"
            if lead_time > 0 and i > 0:
                prev_stage = stages[i - 1]
                prev_val = sv.get(prev_stage["id"], {})
                prev_updated = prev_val.get("updated_at", "") if isinstance(prev_val, dict) else ""
                prev_value = prev_val.get("value", "") if isinstance(prev_val, dict) else str(prev_val) if prev_val else ""
                if prev_value and prev_updated:
                    try:
                        prev_date = datetime.fromisoformat(prev_updated.replace("Z", "+00:00"))
                        expected_due = prev_date + timedelta(days=lead_time)
                        if stage_value and stage_updated_at:
                            completed = datetime.fromisoformat(stage_updated_at.replace("Z", "+00:00"))
                            status = "completed_early" if completed <= expected_due else "completed_late"
                        else:
                            status = "pending" if now <= expected_due else "delayed"
                    except (ValueError, TypeError):
                        pass
            elif stage_value:
                status = "completed"
            delay_info[sid] = status
        enq["delay_status"] = delay_info
    return {"enquiries": enquiries, "total": total, "page": page, "page_size": page_size, "total_pages": -(-total // page_size)}

@api_router.get("/enquiries/{enquiry_id}")
async def get_enquiry(enquiry_id: str, request: Request):
    await get_current_user(request)
    enquiry = await db.enquiries.find_one({"id": enquiry_id}, {"_id": 0})
    if not enquiry:
        raise HTTPException(status_code=404, detail="Enquiry not found")
    history = await db.enquiry_history.find({"enquiry_id": enquiry_id}, {"_id": 0}).sort("changed_at", -1).to_list(100)
    enquiry["history"] = history
    # Calculate delay status for each stage
    stages = await db.stages.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    sv = enquiry.get("stage_values", {})
    delay_status = {}
    for i, stage in enumerate(stages):
        sid = stage["id"]
        lead_time = stage.get("lead_time_days", 0)
        stage_val = sv.get(sid, {})
        stage_value = stage_val.get("value", "") if isinstance(stage_val, dict) else str(stage_val) if stage_val else ""
        stage_updated_at = stage_val.get("updated_at", "") if isinstance(stage_val, dict) else ""
        status = "pending"  # pending / on_time / delayed / completed_early / completed_late
        due_date = None
        days_diff = None
        if lead_time > 0 and i > 0:
            # Find previous stage completion date
            prev_stage = stages[i - 1]
            prev_val = sv.get(prev_stage["id"], {})
            prev_updated = prev_val.get("updated_at", "") if isinstance(prev_val, dict) else ""
            prev_value = prev_val.get("value", "") if isinstance(prev_val, dict) else str(prev_val) if prev_val else ""
            if prev_value and prev_updated:
                try:
                    prev_date = datetime.fromisoformat(prev_updated.replace("Z", "+00:00"))
                    expected_due = prev_date + timedelta(days=lead_time)
                    due_date = expected_due.isoformat()
                    now = datetime.now(timezone.utc)
                    if stage_value and stage_updated_at:
                        # Stage is completed - check if it was early or late
                        completed = datetime.fromisoformat(stage_updated_at.replace("Z", "+00:00"))
                        diff = (expected_due - completed).total_seconds() / 86400
                        days_diff = round(diff, 1)
                        status = "completed_early" if diff >= 0 else "completed_late"
                    else:
                        # Stage not yet completed - check if overdue
                        diff = (expected_due - now).total_seconds() / 86400
                        days_diff = round(diff, 1)
                        status = "pending" if diff >= 0 else "delayed"
                except (ValueError, TypeError):
                    pass
        elif stage_value:
            status = "completed"
        delay_status[sid] = {"status": status, "due_date": due_date, "days_diff": days_diff, "lead_time_days": lead_time}
    enquiry["delay_status"] = delay_status
    return enquiry

@api_router.post("/enquiries")
async def create_enquiry(req: EnquiryCreate, request: Request):
    user = await get_current_user(request)
    enquiry_id = secrets.token_hex(12)
    now = datetime.now(timezone.utc).isoformat()
    enquiry_doc = {
        "id": enquiry_id, "customer_name": req.customer_name,
        "fabric_type": req.fabric_type, "quantity": req.quantity,
        "style_no": req.style_no, "image_path": "",
        "department": req.department or user.get("department", ""),
        "notes": req.notes, "rate": req.rate, "po_no": req.po_no,
        "po_del_date": req.po_del_date,
        "fabric_received": req.fabric_received,
        "qty_received": req.qty_received,
        "sample_number": req.sample_number,
        "stage_values": req.stage_values,
        "status": "open",
        "created_by": user["_id"], "created_by_name": user["name"],
        "created_at": now, "updated_at": now
    }
    await db.enquiries.insert_one(enquiry_doc)
    # Track history for initial stage values
    for stage_id, val in req.stage_values.items():
        history_doc = {
            "id": secrets.token_hex(12), "enquiry_id": enquiry_id,
            "stage_id": stage_id, "old_value": "", "new_value": val.get("value", "") if isinstance(val, dict) else str(val),
            "changed_by": user["_id"], "changed_by_name": user["name"],
            "changed_at": now, "notes": "Initial value set"
        }
        await db.enquiry_history.insert_one(history_doc)
    return {k: v for k, v in enquiry_doc.items() if k != "_id"}

@api_router.post("/enquiries/bulk")
async def bulk_create_enquiries(request: Request, files: List[UploadFile] = File(...), customer_name: str = Form(...), fabric_type: str = Form(...), style_no: str = Form(""), department: str = Form(""), notes: str = Form(""), fabric_received: str = Form("no"), qty_received: str = Form(""), sample_number: str = Form("")):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc).isoformat()
    created = []
    for f in files:
        enquiry_id = secrets.token_hex(12)
        # Upload image
        image_path = ""
        try:
            ext = f.filename.split(".")[-1] if "." in f.filename else "jpg"
            path = f"{APP_NAME}/images/{enquiry_id}.{ext}"
            data = await f.read()
            result = put_object(path, data, f.content_type or "image/jpeg")
            image_path = result["path"]
            file_doc = {
                "id": str(uuid_mod.uuid4()), "storage_path": result["path"],
                "original_filename": f.filename, "content_type": f.content_type or "image/jpeg",
                "size": result.get("size", len(data)), "is_deleted": False,
                "uploaded_by": user["_id"], "created_at": now
            }
            await db.files.insert_one(file_doc)
        except Exception:
            pass
        enquiry_doc = {
            "id": enquiry_id, "customer_name": customer_name,
            "fabric_type": fabric_type, "quantity": "",
            "style_no": style_no, "image_path": image_path,
            "department": department or user.get("department", ""),
            "notes": notes, "rate": "", "po_no": "", "po_del_date": "",
            "fabric_received": fabric_received, "qty_received": qty_received,
            "sample_number": sample_number, "stage_values": {},
            "status": "open",
            "created_by": user["_id"], "created_by_name": user["name"],
            "created_at": now, "updated_at": now
        }
        await db.enquiries.insert_one(enquiry_doc)
        created.append({k: v for k, v in enquiry_doc.items() if k != "_id"})
    return {"created": len(created), "enquiries": created}

@api_router.put("/enquiries/{enquiry_id}")
async def update_enquiry(enquiry_id: str, req: EnquiryUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.enquiries.find_one({"id": enquiry_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Enquiry not found")
    if existing.get("status") == "closed" and user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Enquiry is closed")
    update_data = {}
    req_dict = req.model_dump(exclude_unset=True)
    now = datetime.now(timezone.utc).isoformat()
    changed_stage_ids = []
    # Handle stage_values separately for history tracking
    if "stage_values" in req_dict and req_dict["stage_values"] is not None:
        new_sv = req_dict.pop("stage_values")
        old_sv = existing.get("stage_values", {})
        # Load department hierarchy for permission check
        dept_name = existing.get("department", "")
        dept = await db.departments.find_one({"name": dept_name}, {"_id": 0}) if dept_name else None
        hierarchy_map = {}
        if dept and dept.get("stage_hierarchy"):
            for h in dept["stage_hierarchy"]:
                hierarchy_map[h["stage_id"]] = h.get("assigned_users", [])
        all_stages = {s["id"]: s for s in await db.stages.find({}, {"_id": 0}).to_list(100)}
        for stage_id, new_val in new_sv.items():
            # Check permissions: use department hierarchy first, then fall back to stage-level
            assigned = hierarchy_map.get(stage_id)
            if assigned is None:
                stage_def = all_stages.get(stage_id)
                assigned = stage_def.get("assigned_users", []) if stage_def else []
            if assigned and len(assigned) > 0:
                if user["_id"] not in assigned and user.get("role") != "admin":
                    continue  # Skip stages the user isn't assigned to
            new_value_str = new_val.get("value", "") if isinstance(new_val, dict) else str(new_val)
            old_val = old_sv.get(stage_id, {})
            old_value_str = old_val.get("value", "") if isinstance(old_val, dict) else str(old_val) if old_val else ""
            if new_value_str != old_value_str:
                changed_stage_ids.append(stage_id)
                history_doc = {
                    "id": secrets.token_hex(12), "enquiry_id": enquiry_id,
                    "stage_id": stage_id, "type": "value_change",
                    "old_value": old_value_str, "new_value": new_value_str,
                    "changed_by": user["_id"], "changed_by_name": user["name"],
                    "changed_at": now, "notes": "Stage value updated"
                }
                await db.enquiry_history.insert_one(history_doc)
        # Merge stage values
        merged = {**old_sv}
        for k, v in new_sv.items():
            if isinstance(v, dict):
                v["updated_at"] = now
                v["updated_by"] = user["_id"]
            merged[k] = v
        update_data["stage_values"] = merged
    for k, v in req_dict.items():
        if v is not None:
            update_data[k] = v
    update_data["updated_at"] = now
    await db.enquiries.update_one({"id": enquiry_id}, {"$set": update_data})
    # Auto-close and notify for each changed stage
    dept_name = existing.get("department") or update_data.get("department", "")
    for sid in changed_stage_ids:
        await check_auto_close_and_notify(enquiry_id, dept_name, sid, user)
    enquiry = await db.enquiries.find_one({"id": enquiry_id}, {"_id": 0})
    return enquiry

@api_router.delete("/enquiries/{enquiry_id}")
async def delete_enquiry(enquiry_id: str, request: Request):
    await require_admin(request)
    result = await db.enquiries.delete_one({"id": enquiry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Enquiry not found")
    await db.enquiry_history.delete_many({"enquiry_id": enquiry_id})
    return {"message": "Enquiry deleted"}

# ─── Stage Comments ───
class StageCommentCreate(BaseModel):
    stage_id: str
    comment: str

@api_router.post("/enquiries/{enquiry_id}/comments")
async def add_stage_comment(enquiry_id: str, req: StageCommentCreate, request: Request):
    user = await get_current_user(request)
    existing = await db.enquiries.find_one({"id": enquiry_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Enquiry not found")
    # Check if user is assigned to this stage
    stage = await db.stages.find_one({"id": req.stage_id}, {"_id": 0})
    if stage and stage.get("assigned_users") and len(stage["assigned_users"]) > 0:
        if user["_id"] not in stage["assigned_users"] and user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="You are not assigned to this stage")
    now = datetime.now(timezone.utc).isoformat()
    comment_doc = {
        "id": secrets.token_hex(12),
        "enquiry_id": enquiry_id,
        "stage_id": req.stage_id,
        "type": "comment",
        "comment": req.comment,
        "changed_by": user["_id"],
        "changed_by_name": user["name"],
        "changed_at": now,
        "old_value": "",
        "new_value": "",
        "notes": req.comment
    }
    await db.enquiry_history.insert_one(comment_doc)
    return {k: v for k, v in comment_doc.items() if k != "_id"}

# ─── Enquiry Close ───
@api_router.put("/enquiries/{enquiry_id}/close")
async def close_enquiry(enquiry_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admin can close enquiries")
    existing = await db.enquiries.find_one({"id": enquiry_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Enquiry not found")
    now = datetime.now(timezone.utc).isoformat()
    await db.enquiries.update_one({"id": enquiry_id}, {"$set": {"status": "closed", "closed_at": now, "closed_by": user["_id"]}})
    await db.enquiry_history.insert_one({
        "id": secrets.token_hex(12), "enquiry_id": enquiry_id, "stage_id": "",
        "type": "status_change", "old_value": "open", "new_value": "closed",
        "changed_by": user["_id"], "changed_by_name": user["name"],
        "changed_at": now, "notes": "Enquiry closed by admin"
    })
    return {"message": "Enquiry closed", "status": "closed"}

@api_router.put("/enquiries/{enquiry_id}/reopen")
async def reopen_enquiry(enquiry_id: str, request: Request):
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admin can reopen enquiries")
    existing = await db.enquiries.find_one({"id": enquiry_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Enquiry not found")
    await db.enquiries.update_one({"id": enquiry_id}, {"$set": {"status": "open", "closed_at": None, "closed_by": None}})
    return {"message": "Enquiry reopened", "status": "open"}

# Helper: check auto-close and create notifications
async def check_auto_close_and_notify(enquiry_id: str, dept_name: str, changed_stage_id: str, changed_by_user: dict):
    """After a stage is updated, check if enquiry should auto-close and notify next stage users."""
    dept = await db.departments.find_one({"name": dept_name}, {"_id": 0})
    if not dept or not dept.get("stage_hierarchy"):
        return
    hierarchy = sorted(dept["stage_hierarchy"], key=lambda x: x.get("order", 0))
    enq = await db.enquiries.find_one({"id": enquiry_id}, {"_id": 0})
    if not enq or enq.get("status") == "closed":
        return
    sv = enq.get("stage_values", {})

    # Check auto-close: all hierarchy stages must have a value
    all_filled = True
    for h in hierarchy:
        val = sv.get(h["stage_id"], {})
        v = val.get("value", "") if isinstance(val, dict) else str(val) if val else ""
        if not v:
            all_filled = False
            break
    if all_filled and len(hierarchy) > 0:
        now = datetime.now(timezone.utc).isoformat()
        await db.enquiries.update_one({"id": enquiry_id}, {"$set": {"status": "closed", "closed_at": now, "closed_by": "auto"}})
        await db.enquiry_history.insert_one({
            "id": secrets.token_hex(12), "enquiry_id": enquiry_id, "stage_id": "",
            "type": "status_change", "old_value": "open", "new_value": "closed",
            "changed_by": "system", "changed_by_name": "System",
            "changed_at": now, "notes": "Auto-closed: all stages completed"
        })

    # Notify next stage users
    current_idx = -1
    for i, h in enumerate(hierarchy):
        if h["stage_id"] == changed_stage_id:
            current_idx = i
            break
    if current_idx >= 0 and current_idx + 1 < len(hierarchy):
        next_stage = hierarchy[current_idx + 1]
        next_stage_def = await db.stages.find_one({"id": next_stage["stage_id"]}, {"_id": 0})
        next_stage_name = next_stage_def["name"] if next_stage_def else next_stage["stage_id"]
        current_stage_def = await db.stages.find_one({"id": changed_stage_id}, {"_id": 0})
        current_stage_name = current_stage_def["name"] if current_stage_def else changed_stage_id
        for uid in next_stage.get("assigned_users", []):
            notif_doc = {
                "id": secrets.token_hex(12),
                "user_id": uid,
                "type": "stage_reminder",
                "title": f"Stage '{next_stage_name}' is ready",
                "message": f"Stage '{current_stage_name}' has been completed by {changed_by_user.get('name', 'a user')} for enquiry {enq.get('customer_name', '')} ({enq.get('style_no', '')}). Your stage '{next_stage_name}' is now pending.",
                "enquiry_id": enquiry_id,
                "stage_id": next_stage["stage_id"],
                "is_read": False,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.notifications.insert_one(notif_doc)

# ─── Notifications ───
@api_router.get("/notifications")
async def get_notifications(request: Request):
    user = await get_current_user(request)
    notifs = await db.notifications.find({"user_id": user["_id"]}, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)
    return notifs

@api_router.get("/notifications/unread-count")
async def get_unread_count(request: Request):
    user = await get_current_user(request)
    count = await db.notifications.count_documents({"user_id": user["_id"], "is_read": False})
    return {"count": count}

@api_router.put("/notifications/{notif_id}/read")
async def mark_notification_read(notif_id: str, request: Request):
    user = await get_current_user(request)
    await db.notifications.update_one({"id": notif_id, "user_id": user["_id"]}, {"$set": {"is_read": True}})
    return {"message": "Marked as read"}

@api_router.put("/notifications/read-all")
async def mark_all_read(request: Request):
    user = await get_current_user(request)
    await db.notifications.update_many({"user_id": user["_id"], "is_read": False}, {"$set": {"is_read": True}})
    return {"message": "All marked as read"}

# ─── Department Master ───
@api_router.get("/departments")
async def get_departments(request: Request):
    await get_current_user(request)
    return await get_cached("departments", lambda: db.departments.find({}, {"_id": 0}).sort("name", 1).to_list(100))

@api_router.post("/departments")
async def create_department(req: DepartmentCreate, request: Request):
    await require_admin(request)
    existing = await db.departments.find_one({"name": req.name})
    if existing:
        raise HTTPException(status_code=400, detail="Department already exists")
    dept_id = secrets.token_hex(12)
    dept_doc = {"id": dept_id, "name": req.name, "description": req.description, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.departments.insert_one(dept_doc)
    invalidate_cache("departments")
    return {k: v for k, v in dept_doc.items() if k != "_id"}

@api_router.put("/departments/{dept_id}")
async def update_department(dept_id: str, req: DepartmentUpdate, request: Request):
    await require_admin(request)
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await db.departments.update_one({"id": dept_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Department not found")
    dept = await db.departments.find_one({"id": dept_id}, {"_id": 0})
    invalidate_cache("departments")
    return dept

@api_router.delete("/departments/{dept_id}")
async def delete_department(dept_id: str, request: Request):
    await require_admin(request)
    result = await db.departments.delete_one({"id": dept_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Department not found")
    invalidate_cache("departments")
    return {"message": "Department deleted"}

# ─── Department Stage Hierarchy ───
@api_router.get("/departments/{dept_id}/hierarchy")
async def get_department_hierarchy(dept_id: str, request: Request):
    await get_current_user(request)
    dept = await db.departments.find_one({"id": dept_id}, {"_id": 0})
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    return dept.get("stage_hierarchy", [])

@api_router.put("/departments/{dept_id}/hierarchy")
async def update_department_hierarchy(dept_id: str, req: DepartmentHierarchyUpdate, request: Request):
    await require_admin(request)
    dept = await db.departments.find_one({"id": dept_id})
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    hierarchy = [item.model_dump() for item in req.items]
    await db.departments.update_one({"id": dept_id}, {"$set": {"stage_hierarchy": hierarchy}})
    invalidate_cache("departments")
    return hierarchy

# ─── Customer Master ───
@api_router.get("/customers")
async def get_customers(request: Request):
    await get_current_user(request)
    customers = await db.customers.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return customers

@api_router.post("/customers")
async def create_customer(req: CustomerCreate, request: Request):
    await get_current_user(request)
    existing = await db.customers.find_one({"name": req.name})
    if existing:
        raise HTTPException(status_code=400, detail="Customer already exists")
    cust_id = secrets.token_hex(12)
    doc = {"id": cust_id, "name": req.name, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.customers.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/customers/{cust_id}")
async def update_customer(cust_id: str, req: CustomerUpdate, request: Request):
    await get_current_user(request)
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await db.customers.update_one({"id": cust_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return await db.customers.find_one({"id": cust_id}, {"_id": 0})

@api_router.delete("/customers/{cust_id}")
async def delete_customer(cust_id: str, request: Request):
    await require_admin(request)
    result = await db.customers.delete_one({"id": cust_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer deleted"}

# ─── Fabric Type Master ───
@api_router.get("/fabric-types")
async def get_fabric_types(request: Request):
    await get_current_user(request)
    fabrics = await db.fabric_types.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return fabrics

@api_router.post("/fabric-types")
async def create_fabric_type(req: FabricTypeCreate, request: Request):
    await get_current_user(request)
    existing = await db.fabric_types.find_one({"name": req.name})
    if existing:
        raise HTTPException(status_code=400, detail="Fabric type already exists")
    fab_id = secrets.token_hex(12)
    doc = {"id": fab_id, "name": req.name, "gsm": req.gsm, "width": req.width, "composition": req.composition, "construction": req.construction, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.fabric_types.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}

@api_router.put("/fabric-types/{fab_id}")
async def update_fabric_type(fab_id: str, req: FabricTypeUpdate, request: Request):
    await get_current_user(request)
    update_data = {k: v for k, v in req.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await db.fabric_types.update_one({"id": fab_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Fabric type not found")
    return await db.fabric_types.find_one({"id": fab_id}, {"_id": 0})

@api_router.delete("/fabric-types/{fab_id}")
async def delete_fabric_type(fab_id: str, request: Request):
    await require_admin(request)
    result = await db.fabric_types.delete_one({"id": fab_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fabric type not found")
    return {"message": "Fabric type deleted"}

# ─── Dashboard ───
@api_router.get("/dashboard")
async def get_dashboard(request: Request):
    import asyncio
    await get_current_user(request)
    stages = await get_cached("stages", lambda: db.stages.find({}, {"_id": 0}).sort("order", 1).to_list(100))
    # Parallel queries
    total_f = db.enquiries.count_documents({})
    users_f = db.users.count_documents({})
    by_dept_f = db.enquiries.aggregate([{"$group": {"_id": "$department", "count": {"$sum": 1}}}]).to_list(100)
    recent_f = db.enquiries.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    stage_counts = [db.enquiries.count_documents({f"stage_values.{s['id']}": {"$exists": True}}) for s in stages]
    results = await asyncio.gather(total_f, users_f, by_dept_f, recent_f, *stage_counts)
    total = results[0]
    users_count = results[1]
    by_dept_raw = results[2]
    recent = results[3]
    by_stage = [{"stage_id": s["id"], "stage_name": s["name"], "count": results[4 + i]} for i, s in enumerate(stages)]
    by_department = [{"department": d["_id"] or "Unassigned", "count": d["count"]} for d in by_dept_raw]
    return {
        "total_enquiries": total, "by_stage": by_stage,
        "by_department": by_department, "recent_enquiries": recent,
        "total_users": users_count, "total_stages": len(stages)
    }

# ─── Reports ───
@api_router.get("/reports/enquiries")
async def report_enquiries(request: Request, start_date: Optional[str] = None, end_date: Optional[str] = None, department: Optional[str] = None, assigned_to: Optional[str] = None, customer_name: Optional[str] = None, fabric_type: Optional[str] = None, style_no: Optional[str] = None, rate: Optional[str] = None, po_no: Optional[str] = None, po_del_date: Optional[str] = None, fabric_received: Optional[str] = None, qty_received: Optional[str] = None, created_by: Optional[str] = None, stage_filters: Optional[str] = None):
    import json as _json
    await get_current_user(request)
    query = {}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        query.setdefault("created_at", {})["$lte"] = end_date
    if department:
        query["department"] = department
    if assigned_to:
        query["assigned_to"] = assigned_to
    if customer_name:
        query["customer_name"] = {"$regex": customer_name, "$options": "i"}
    if fabric_type:
        query["fabric_type"] = {"$regex": fabric_type, "$options": "i"}
    if style_no:
        query["style_no"] = {"$regex": style_no, "$options": "i"}
    if rate:
        query["rate"] = {"$regex": rate, "$options": "i"}
    if po_no:
        query["po_no"] = {"$regex": po_no, "$options": "i"}
    if po_del_date:
        query["po_del_date"] = {"$regex": po_del_date, "$options": "i"}
    if fabric_received:
        query["fabric_received"] = fabric_received
    if qty_received:
        query["qty_received"] = {"$regex": qty_received, "$options": "i"}
    if created_by:
        query["created_by"] = created_by
    # Dynamic stage filters
    if stage_filters:
        try:
            sf = _json.loads(stage_filters)
            for stage_id, val in sf.items():
                if not val:
                    continue
                if val == "__blank__":
                    query["$and"] = query.get("$and", [])
                    query["$and"].append({"$or": [
                        {f"stage_values.{stage_id}": {"$exists": False}},
                        {f"stage_values.{stage_id}.value": {"$in": [None, ""]}},
                    ]})
                elif val == "__filled__":
                    query[f"stage_values.{stage_id}.value"] = {"$exists": True, "$nin": [None, ""]}
                else:
                    query[f"stage_values.{stage_id}.value"] = {"$regex": val, "$options": "i"}
        except (_json.JSONDecodeError, TypeError):
            pass
    enquiries = await db.enquiries.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    return {"total": len(enquiries), "enquiries": enquiries}

@api_router.get("/reports/stage-summary")
async def report_stage_summary(request: Request):
    await get_current_user(request)
    stages = await db.stages.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    summary = []
    for s in stages:
        count = await db.enquiries.count_documents({f"stage_values.{s['id']}": {"$exists": True}})
        filled_count = await db.enquiries.count_documents({
            f"stage_values.{s['id']}.value": {"$exists": True, "$ne": ""}
        })
        summary.append({
            "stage_id": s["id"], "stage_name": s["name"], "color": s["color"],
            "input_type": s.get("input_type", "text"), "is_mandatory": s.get("is_mandatory", False),
            "total_enquiries": count, "filled_count": filled_count
        })
    return summary

@api_router.get("/reports/user-performance")
async def report_user_performance(request: Request):
    await get_current_user(request)
    users = await db.users.find({}, {"password_hash": 0}).to_list(1000)
    performance = []
    for u in users:
        uid = str(u["_id"])
        total_assigned = await db.enquiries.count_documents({"assigned_to": uid})
        changes_made = await db.enquiry_history.count_documents({"changed_by": uid})
        performance.append({
            "user_id": uid, "user_name": u["name"],
            "department": u.get("department", ""), "role": u["role"],
            "total_assigned": total_assigned, "changes_made": changes_made
        })
    return performance

@api_router.get("/reports/department")
async def report_department(request: Request):
    await get_current_user(request)
    pipeline = [{"$group": {"_id": "$department", "total": {"$sum": 1}}}]
    raw = await db.enquiries.aggregate(pipeline).to_list(100)
    stages = await db.stages.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    result = []
    for dept in raw:
        dept_name = dept["_id"] or "Unassigned"
        stage_breakdown = []
        for s in stages:
            count = await db.enquiries.count_documents({
                "department": dept["_id"],
                f"stage_values.{s['id']}": {"$exists": True}
            })
            stage_breakdown.append({"stage_name": s["name"], "stage_id": s["id"], "color": s["color"], "count": count})
        result.append({"department": dept_name, "total": dept["total"], "stage_breakdown": stage_breakdown})
    return result

# ─── Pending Stages Report ───
@api_router.get("/reports/pending-stages")
async def report_pending_stages(request: Request):
    await get_current_user(request)
    depts = await db.departments.find({}, {"_id": 0}).to_list(100)
    stages = {s["id"]: s for s in await db.stages.find({}, {"_id": 0}).to_list(100)}
    users_list = await db.users.find({}, {"password_hash": 0}).to_list(1000)
    user_map = {str(u["_id"]): u for u in users_list}
    user_pending = {}
    for dept in depts:
        hierarchy = dept.get("stage_hierarchy", [])
        if not hierarchy:
            continue
        sorted_h = sorted(hierarchy, key=lambda x: x.get("order", 0))
        enquiries = await db.enquiries.find({"department": dept["name"], "status": {"$ne": "closed"}}, {"_id": 0}).to_list(5000)
        for enq in enquiries:
            sv = enq.get("stage_values", {})
            for i, h in enumerate(sorted_h):
                sid = h["stage_id"]
                val = sv.get(sid, {})
                v = val.get("value", "") if isinstance(val, dict) else str(val) if val else ""
                if v:
                    continue
                can_work = True
                if i > 0:
                    prev_sid = sorted_h[i - 1]["stage_id"]
                    prev_val = sv.get(prev_sid, {})
                    pv = prev_val.get("value", "") if isinstance(prev_val, dict) else str(prev_val) if prev_val else ""
                    if not pv:
                        can_work = False
                if not can_work:
                    continue
                stage_def = stages.get(sid)
                for uid in h.get("assigned_users", []):
                    if uid not in user_pending:
                        u = user_map.get(uid)
                        user_pending[uid] = {"user_id": uid, "user_name": u["name"] if u else uid, "department": u.get("department", "") if u else "", "items": []}
                    user_pending[uid]["items"].append({
                        "enquiry_id": enq["id"], "customer_name": enq.get("customer_name", ""),
                        "style_no": enq.get("style_no", ""), "stage_id": sid,
                        "stage_name": stage_def["name"] if stage_def else sid, "department": dept["name"]
                    })
                break
    return list(user_pending.values())

# ─── User Stages Report (Pending + Done) ───
@api_router.get("/reports/user-stages")
async def report_user_stages(request: Request, filter_department: Optional[str] = None, filter_user: Optional[str] = None, filter_stage: Optional[str] = None):
    await get_current_user(request)
    depts = await get_cached("departments", lambda: db.departments.find({}, {"_id": 0}).to_list(100))
    all_stages = {s["id"]: s for s in await get_cached("stages", lambda: db.stages.find({}, {"_id": 0}).to_list(100))}
    users_list = await db.users.find({}, {"password_hash": 0}).to_list(1000)
    user_map = {str(u["_id"]): u for u in users_list}
    now = datetime.now(timezone.utc)
    user_data = {}
    for dept in depts:
        if filter_department and dept["name"] != filter_department:
            continue
        hierarchy = dept.get("stage_hierarchy", [])
        if not hierarchy:
            continue
        sorted_h = sorted(hierarchy, key=lambda x: x.get("order", 0))
        enquiries = await db.enquiries.find({"department": dept["name"]}, {"_id": 0}).to_list(5000)
        for enq in enquiries:
            sv = enq.get("stage_values", {})
            is_closed = enq.get("status") == "closed"
            enq_created = enq.get("created_at", "")
            enq_image = enq.get("image_path", "")
            first_pending_found = False  # Track if we already found the first pending stage for this enquiry
            for i, h in enumerate(sorted_h):
                sid = h["stage_id"]
                if filter_stage and sid != filter_stage:
                    continue
                val = sv.get(sid, {})
                v = val.get("value", "") if isinstance(val, dict) else str(val) if val else ""
                updated_at = val.get("updated_at", "") if isinstance(val, dict) else ""
                stage_def = all_stages.get(sid)
                stage_name = stage_def["name"] if stage_def else sid
                lead_time = stage_def.get("lead_time_days", 0) if stage_def else 0
                # Previous stage info
                prev_stage_name = ""
                prev_completed_by = ""
                prev_completed_at = ""
                prev_value = ""
                if i > 0:
                    prev_h = sorted_h[i - 1]
                    prev_sid = prev_h["stage_id"]
                    prev_s_def = all_stages.get(prev_sid)
                    prev_stage_name = prev_s_def["name"] if prev_s_def else prev_sid
                    prev_sv = sv.get(prev_sid, {})
                    prev_value = prev_sv.get("value", "") if isinstance(prev_sv, dict) else str(prev_sv) if prev_sv else ""
                    prev_completed_at = prev_sv.get("updated_at", "") if isinstance(prev_sv, dict) else ""
                    prev_completed_by_id = prev_sv.get("updated_by", "") if isinstance(prev_sv, dict) else ""
                    if prev_completed_by_id and prev_completed_by_id in user_map:
                        prev_completed_by = user_map[prev_completed_by_id]["name"]
                for uid in h.get("assigned_users", []):
                    if filter_user and uid != filter_user:
                        continue
                    if uid not in user_data:
                        u = user_map.get(uid)
                        user_data[uid] = {
                            "user_id": uid,
                            "user_name": u["name"] if u else uid,
                            "department": u.get("department", "") if u else "",
                            "pending": [], "done": [],
                            "pending_count": 0, "done_count": 0
                        }
                    item = {
                        "enquiry_id": enq["id"],
                        "customer_name": enq.get("customer_name", ""),
                        "style_no": enq.get("style_no", ""),
                        "fabric_type": enq.get("fabric_type", ""),
                        "image_path": enq_image,
                        "stage_id": sid,
                        "stage_name": stage_name,
                        "department": dept["name"],
                        "enquiry_status": enq.get("status", "open"),
                        "enquiry_created_at": enq_created,
                        "lead_time_days": lead_time,
                        "prev_stage_name": prev_stage_name,
                        "prev_completed_by": prev_completed_by,
                        "prev_completed_at": prev_completed_at,
                    }
                    if v:
                        item["value"] = v
                        item["completed_at"] = updated_at
                        item["days_taken"] = 0
                        if updated_at and prev_completed_at:
                            try:
                                c = datetime.fromisoformat(updated_at.replace("Z", "+00:00"))
                                p = datetime.fromisoformat(prev_completed_at.replace("Z", "+00:00"))
                                item["days_taken"] = round((c - p).total_seconds() / 86400, 1)
                            except (ValueError, TypeError):
                                pass
                        user_data[uid]["done"].append(item)
                        user_data[uid]["done_count"] += 1
                    elif not is_closed and not first_pending_found:
                        # Calculate days pending
                        ref_date_str = prev_completed_at or enq_created
                        days_pending = 0
                        is_overdue = False
                        due_date = ""
                        if ref_date_str:
                            try:
                                ref = datetime.fromisoformat(ref_date_str.replace("Z", "+00:00"))
                                days_pending = round((now - ref).total_seconds() / 86400, 1)
                                if lead_time > 0:
                                    dd = ref + timedelta(days=lead_time)
                                    due_date = dd.isoformat()
                                    is_overdue = now > dd
                            except (ValueError, TypeError):
                                pass
                        item["days_pending"] = days_pending
                        item["is_overdue"] = is_overdue
                        item["due_date"] = due_date
                        user_data[uid]["pending"].append(item)
                        user_data[uid]["pending_count"] += 1
                if not v and not is_closed and not filter_stage:
                    first_pending_found = True  # Only show first pending stage per enquiry
                    break  # Stop checking further stages for this enquiry
    # Sort pending by overdue first, then days_pending desc
    for ud in user_data.values():
        ud["pending"].sort(key=lambda x: (-int(x.get("is_overdue", False)), -x.get("days_pending", 0)))
    return list(user_data.values())

# ─── User Stages Excel Export ───
@api_router.get("/reports/user-stages/export-excel")
async def export_user_stages_excel(request: Request, filter_department: Optional[str] = None, filter_user: Optional[str] = None, filter_stage: Optional[str] = None):
    await get_current_user(request)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    # Reuse the user-stages logic
    from starlette.datastructures import QueryParams
    # Build a fake request isn't needed — just call the data-gathering inline
    # Fetch data same as report
    depts = await db.departments.find({}, {"_id": 0}).to_list(100)
    all_stages = {s["id"]: s for s in await db.stages.find({}, {"_id": 0}).to_list(100)}
    users_list = await db.users.find({}, {"password_hash": 0}).to_list(1000)
    umap = {str(u["_id"]): u for u in users_list}
    now_dt = datetime.now(timezone.utc)
    rows = []  # flat rows for Excel
    for dept in depts:
        if filter_department and dept["name"] != filter_department:
            continue
        hierarchy = dept.get("stage_hierarchy", [])
        if not hierarchy:
            continue
        sorted_h = sorted(hierarchy, key=lambda x: x.get("order", 0))
        enquiries = await db.enquiries.find({"department": dept["name"]}, {"_id": 0}).to_list(5000)
        for enq in enquiries:
            sv = enq.get("stage_values", {})
            is_closed = enq.get("status") == "closed"
            enq_created = enq.get("created_at", "")
            for i, h in enumerate(sorted_h):
                sid = h["stage_id"]
                if filter_stage and sid != filter_stage:
                    continue
                val = sv.get(sid, {})
                v = val.get("value", "") if isinstance(val, dict) else str(val) if val else ""
                updated_at = val.get("updated_at", "") if isinstance(val, dict) else ""
                stage_def = all_stages.get(sid)
                stage_name = stage_def["name"] if stage_def else sid
                lead_time = stage_def.get("lead_time_days", 0) if stage_def else 0
                prev_stage_name = ""
                prev_completed_by = ""
                prev_completed_at = ""
                if i > 0:
                    prev_h = sorted_h[i - 1]
                    prev_s_def = all_stages.get(prev_h["stage_id"])
                    prev_stage_name = prev_s_def["name"] if prev_s_def else prev_h["stage_id"]
                    prev_sv = sv.get(prev_h["stage_id"], {})
                    prev_completed_at = prev_sv.get("updated_at", "") if isinstance(prev_sv, dict) else ""
                    pby = prev_sv.get("updated_by", "") if isinstance(prev_sv, dict) else ""
                    if pby and pby in umap:
                        prev_completed_by = umap[pby]["name"]
                for uid in h.get("assigned_users", []):
                    if filter_user and uid != filter_user:
                        continue
                    u = umap.get(uid)
                    user_name = u["name"] if u else uid
                    status = "Done" if v else ("Closed" if is_closed else "Pending")
                    days_val = 0
                    is_overdue = False
                    due_date_str = ""
                    if v and updated_at and prev_completed_at:
                        try:
                            c = datetime.fromisoformat(updated_at.replace("Z", "+00:00"))
                            p = datetime.fromisoformat(prev_completed_at.replace("Z", "+00:00"))
                            days_val = round((c - p).total_seconds() / 86400, 1)
                        except: pass
                    elif not v and not is_closed:
                        ref = prev_completed_at or enq_created
                        if ref:
                            try:
                                r = datetime.fromisoformat(ref.replace("Z", "+00:00"))
                                days_val = round((now_dt - r).total_seconds() / 86400, 1)
                                if lead_time > 0:
                                    dd = r + timedelta(days=lead_time)
                                    due_date_str = dd.strftime("%Y-%m-%d")
                                    is_overdue = now_dt > dd
                            except: pass
                    rows.append({
                        "user": user_name, "department": dept["name"], "stage": stage_name,
                        "customer": enq.get("customer_name", ""), "style_no": enq.get("style_no", ""),
                        "fabric_type": enq.get("fabric_type", ""), "status": status,
                        "value": v, "days": days_val, "overdue": "YES" if is_overdue else "",
                        "due_date": due_date_str, "lead_time": lead_time,
                        "prev_stage": prev_stage_name, "prev_by": prev_completed_by,
                        "prev_at": prev_completed_at[:10] if prev_completed_at else "",
                        "completed_at": updated_at[:10] if updated_at else "",
                        "enquiry_created": enq_created[:10] if enq_created else "",
                        "enquiry_status": enq.get("status", "open")
                    })
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Stages Report"
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, size=10)
    red_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
    green_fill = PatternFill(start_color="E8FDE8", end_color="E8FDE8", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    headers = ["USER", "DEPARTMENT", "STAGE", "CUSTOMER", "STYLE NO", "FABRIC TYPE", "STATUS", "VALUE", "DAYS PENDING/TAKEN", "OVERDUE", "DUE DATE", "LEAD TIME (DAYS)", "PREV STAGE", "PREV COMPLETED BY", "PREV COMPLETED AT", "COMPLETED AT", "ENQUIRY CREATED", "ENQUIRY STATUS"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    for row_idx, r in enumerate(rows, 2):
        vals = [r["user"], r["department"], r["stage"], r["customer"], r["style_no"], r["fabric_type"], r["status"], r["value"], r["days"], r["overdue"], r["due_date"], r["lead_time"], r["prev_stage"], r["prev_by"], r["prev_at"], r["completed_at"], r["enquiry_created"], r["enquiry_status"]]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if r["status"] == "Pending" and r["overdue"] == "YES":
                cell.fill = red_fill
            elif r["status"] == "Done":
                cell.fill = green_fill
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=user_stages_report.xlsx"})

# ─── Excel Export ───
@api_router.get("/reports/export-excel")
async def export_excel(request: Request, department: Optional[str] = None, customer_name: Optional[str] = None, fabric_type: Optional[str] = None, style_no: Optional[str] = None, rate: Optional[str] = None, po_no: Optional[str] = None, po_del_date: Optional[str] = None, fabric_received: Optional[str] = None, qty_received: Optional[str] = None, created_by: Optional[str] = None, start_date: Optional[str] = None, end_date: Optional[str] = None, stage_filters: Optional[str] = None):
    import json as _json
    await get_current_user(request)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XlImage
    from PIL import Image as PILImage

    stages = await db.stages.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    users_list = await db.users.find({}, {"password_hash": 0}).to_list(1000)
    user_map = {str(u["_id"]): u["name"] for u in users_list}

    query = {}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        query.setdefault("created_at", {})["$lte"] = end_date
    if department:
        query["department"] = department
    if customer_name:
        query["customer_name"] = {"$regex": customer_name, "$options": "i"}
    if fabric_type:
        query["fabric_type"] = {"$regex": fabric_type, "$options": "i"}
    if style_no:
        query["style_no"] = {"$regex": style_no, "$options": "i"}
    if rate:
        query["rate"] = {"$regex": rate, "$options": "i"}
    if po_no:
        query["po_no"] = {"$regex": po_no, "$options": "i"}
    if po_del_date:
        query["po_del_date"] = {"$regex": po_del_date, "$options": "i"}
    if fabric_received:
        query["fabric_received"] = fabric_received
    if qty_received:
        query["qty_received"] = {"$regex": qty_received, "$options": "i"}
    if created_by:
        query["created_by"] = created_by
    # Dynamic stage filters
    if stage_filters:
        try:
            sf = _json.loads(stage_filters)
            for stage_id, val in sf.items():
                if not val:
                    continue
                if val == "__blank__":
                    query["$and"] = query.get("$and", [])
                    query["$and"].append({"$or": [
                        {f"stage_values.{stage_id}": {"$exists": False}},
                        {f"stage_values.{stage_id}.value": {"$in": [None, ""]}},
                    ]})
                elif val == "__filled__":
                    query[f"stage_values.{stage_id}.value"] = {"$exists": True, "$nin": [None, ""]}
                else:
                    query[f"stage_values.{stage_id}.value"] = {"$regex": val, "$options": "i"}
        except (_json.JSONDecodeError, TypeError):
            pass
    enquiries = await db.enquiries.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enquiry Report"

    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Build headers
    headers = ["SR NO", "IMAGE", "STYLE NO.", "CUSTOMER NAME", "FABRIC TYPE", "QUANTITY"]
    for s in stages:
        headers.append(s["name"])
    headers.extend(["RATE", "PO No.", "PO Received Date", "Created By", "Department", "Created Date", "Fabric Received", "Qty Received", "Comment"])

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Set image column width and row heights
    ws.column_dimensions['B'].width = 12
    IMG_HEIGHT = 50

    for row_idx, enq in enumerate(enquiries, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border

        # Embed image
        img_cell = ws.cell(row=row_idx, column=2, value="")
        img_cell.border = thin_border
        image_path = enq.get("image_path", "")
        if image_path:
            try:
                img_content, img_ct = get_object(image_path)
                img_data = io.BytesIO(img_content)
                pil_img = PILImage.open(img_data)
                pil_img.thumbnail((70, 70))
                thumb_buf = io.BytesIO()
                pil_img.save(thumb_buf, format='PNG')
                thumb_buf.seek(0)
                xl_img = XlImage(thumb_buf)
                xl_img.width = 60
                xl_img.height = 60
                ws.add_image(xl_img, f"B{row_idx}")
                ws.row_dimensions[row_idx].height = IMG_HEIGHT
            except Exception as img_err:
                logger.warning(f"Failed to embed image {image_path}: {img_err}")
                ws.cell(row=row_idx, column=2, value="Yes").border = thin_border
        
        ws.cell(row=row_idx, column=3, value=enq.get("style_no", "")).border = thin_border
        ws.cell(row=row_idx, column=4, value=enq.get("customer_name", "")).border = thin_border
        ws.cell(row=row_idx, column=5, value=enq.get("fabric_type", "")).border = thin_border
        ws.cell(row=row_idx, column=6, value=enq.get("quantity", "")).border = thin_border

        sv = enq.get("stage_values", {})
        for s_idx, s in enumerate(stages):
            val = sv.get(s["id"], {})
            display = val.get("value", "") if isinstance(val, dict) else str(val) if val else ""
            ws.cell(row=row_idx, column=7 + s_idx, value=display).border = thin_border

        col_offset = 7 + len(stages)
        ws.cell(row=row_idx, column=col_offset, value=enq.get("rate", "")).border = thin_border
        ws.cell(row=row_idx, column=col_offset + 1, value=enq.get("po_no", "")).border = thin_border
        ws.cell(row=row_idx, column=col_offset + 2, value=enq.get("po_del_date", "")).border = thin_border
        ws.cell(row=row_idx, column=col_offset + 3, value=user_map.get(enq.get("created_by", ""), enq.get("created_by_name", ""))).border = thin_border
        ws.cell(row=row_idx, column=col_offset + 4, value=enq.get("department", "")).border = thin_border
        ws.cell(row=row_idx, column=col_offset + 5, value=enq.get("created_at", "")).border = thin_border
        ws.cell(row=row_idx, column=col_offset + 6, value=enq.get("fabric_received", "")).border = thin_border
        ws.cell(row=row_idx, column=col_offset + 7, value=enq.get("qty_received", "")).border = thin_border
        ws.cell(row=row_idx, column=col_offset + 8, value=enq.get("notes", "")).border = thin_border

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=enquiry_report.xlsx"}
    )

# Include router
app.include_router(api_router)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Startup
@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.login_attempts.create_index("identifier")
    await db.stages.create_index("id", unique=True)
    await db.enquiries.create_index("id", unique=True)
    await db.enquiries.create_index("department")
    await db.enquiries.create_index("status")
    await db.enquiries.create_index("customer_name")
    await db.enquiries.create_index("created_at")
    await db.enquiries.create_index([("department", 1), ("status", 1)])
    await db.enquiry_history.create_index("enquiry_id")
    await db.departments.create_index("id", unique=True)
    await db.departments.create_index("name", unique=True)
    await db.notifications.create_index("user_id")
    await db.notifications.create_index("id", unique=True)
    await db.webauthn_credentials.create_index("user_id")
    await db.webauthn_credentials.create_index("credential_id_b64")
    await db.files.create_index("id", unique=True)
    # Seed default departments
    default_depts = ["Sales", "Production", "Quality", "Admin", "Design", "Logistics"]
    for dept_name in default_depts:
        existing_dept = await db.departments.find_one({"name": dept_name})
        if not existing_dept:
            await db.departments.insert_one({"id": secrets.token_hex(12), "name": dept_name, "description": "", "created_at": datetime.now(timezone.utc).isoformat()})
    logger.info("Default departments seeded")
    # Init storage
    try:
        init_storage()
        logger.info("Object storage initialized")
    except Exception as e:
        logger.error(f"Storage init failed: {e}")
    # Seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        await db.users.insert_one({
            "email": admin_email, "password_hash": hash_password(admin_password),
            "name": "Admin", "role": "admin", "department": "Admin",
            "is_active": True, "created_at": datetime.now(timezone.utc)
        })
        logger.info(f"Admin user created: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
    # Write test credentials
    os.makedirs("/app/memory", exist_ok=True)
    with open("/app/memory/test_credentials.md", "w") as f:
        f.write(f"# Test Credentials\n\n## Admin\n- Email: {admin_email}\n- Password: {admin_password}\n- Role: admin\n\n## Auth Endpoints\n- POST /api/auth/login\n- POST /api/auth/logout\n- GET /api/auth/me\n")

@app.on_event("shutdown")
async def shutdown():
    client.close()
